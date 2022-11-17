#!/usr/bin/env python
# -*- coding: UTF-8 -*-
"""
@Project ：FinanceTool
@File    ：ReadExcel.py
@Author  ：朱桃禾 MikePy
@Date    ：2022/10/8 10:47 AM
"""

import openpyxl
from DataModule import BasicMessage
from openpyxl.utils import get_column_letter


class FirstDeal:
    def __init__(self, filepath):
        self.file_path = filepath
        self.company_sheet_detail = {}  # 所有数据
        self.time_data = []  # 可选日期
        self.txt_file = open("log.txt", "w", encoding="utf-8")
        # 尝试打开文件，得到wb
        try:
            self.wb = openpyxl.load_workbook(filepath, data_only=True)
            self.workbook_sheets_names = self.wb.sheetnames
        except:
            raise "Can't Open Excel File"
        self.deal_sheets()
        self.get_time_data()
        self.calculate_year_basis()
        BasicMessage.time_data = self.time_data

    def deal_sheets(self):
        """
        处理文件相关sheets name 数据规整
        company_sheet_detail  :  各公司简写名字的对应有的表名
        :return: self.company_sheet_detail
        """

        # 公司简写
        company_abb = list(BasicMessage.company_abbreviation.values())
        # company_all = list(BasicMessage.company_abbreviation.keys())  #全写

        # 构建公司字典
        for x in company_abb:
            if x not in self.company_sheet_detail:
                self.company_sheet_detail[x] = {}
            else:
                pass
        self.txt_file.write("公司字典组合成功！")
        # 读取所有表格中所有的信息，并按照预定格式组合
        for item_company_abb in company_abb:  #
            for item_sheet_name in self.workbook_sheets_names:
                if item_company_abb in item_sheet_name:  # 找到对应表
                    # 读取单张表格 获取未经处理的所有数据
                    self.txt_file.write(f"尝试获取{item_sheet_name}表相关数据！")
                    sheet_data = self.read_single_sheet_table(sheet_name=item_sheet_name)
                    self.txt_file.write(f"获取{item_sheet_name}表相关数据！成功！")
                    # 处理数据，得到组合数据
                    self.txt_file.write(f"开始重建{item_sheet_name}表相关数据！")
                    finally_sheet_data = self.rebuilt_page_data(sheet_data)
                    self.txt_file.write(f"重建{item_sheet_name}表相关数据！成功！")
                    # 组合数据
                    self.company_sheet_detail[item_company_abb][item_sheet_name] = finally_sheet_data
                    self.txt_file.write(f"组合数据成功")
        # 可接受返回的数据，也可以直接调取属性
        return self.company_sheet_detail

    def read_single_sheet_table(self, sheet_name):
        """
        获取单张表格的所有数据
        :param sheet_name:
        :return: data
        """
        ws = self.wb[sheet_name]
        max_row_num = ws.max_row
        max_col_num = ws.max_column
        data = []
        for row in range(2, max_row_num + 1):
            row_item = []
            for col in range(1, max_col_num):
                col_letter = get_column_letter(col)
                row_item.append(ws[col_letter + str(row)].value)
            data.append(row_item)
        return data

    def rebuilt_page_data(self, data_list):
        """
        组装单张表格里面的所有数据
        page_data = {
            "Level1": {"code": {
                                "部门":"code"
                                },
                        "data": [
                                    ["科目代码","部门","项目","本期借方发生","本期借方累积"],
                                ]
                    },
            "Level2": {"code": [], "data": []},
            "Level3": {"code": [], "data": []},
            "Level4": {"code": [], "data": []},
            "Level5": {"code": [], "data": []},
             }
        :return: page_data
        """
        # 预设数据
        page_data = {
            "Level1": {"code": [], "data": [], "total": {}},
            "Level2": {"code": [], "data": [], "total": {}},
            "Level3": {"code": [], "data": [], "total": {}},
            "Level4": {"code": [], "data": [], "total": {}},
            "Level5": {"code": [], "data": [], "total": {}},
        }
        # 有效一级标题
        useful_Level_1 = ["6601", "6602", "6603"]

        for data_index in data_list:
            # 遍历所有数据
            if data_index is None or data_index[0] is None:
                # 无效数据第一批次就跳过
                pass
            else:
                for useful_Level in useful_Level_1:
                    # 只针对有效一级标题下的所有数据
                    if useful_Level in data_index[0]:
                        # 有效数据
                        # ["科目代码","部门","项目","本期借方发生","本期借方累积"]
                        data_detail = [data_index[0], data_index[1], data_index[2], data_index[8], data_index[10]]
                        # 组织一级标题及数据
                        if len(data_index[0]) == 4 and data_index[2] is None:
                            page_data["Level1"]["code"].append(data_index[0] + "-" + data_index[1])
                            page_data["Level1"]["total"][data_index[0] + "-" + data_index[1]] = [data_index[8], data_index[10]]
                        elif len(data_index[0]) == 4 and data_index[2] is not None:
                            page_data["Level1"]["data"].append(data_detail)

                        # 组织二级
                        elif len(data_index[0]) == 7 and data_index[2] is None:
                            page_data["Level2"]["code"].append(data_index[0] + "-" + data_index[1])
                            page_data["Level2"]["total"][data_index[0] + "-" + data_index[1]] = [data_index[8],
                                                                                                 data_index[10]]
                        elif len(data_index[0]) == 7 and data_index[2] is not None:
                            page_data["Level2"]["data"].append(data_detail)

                        # 组织三级
                        elif len(data_index[0]) == 10 and data_index[2] is None:
                            page_data["Level3"]["code"].append(data_index[0] + "-" + data_index[1])
                            page_data["Level3"]["total"][data_index[0] + "-" + data_index[1]] = [data_index[8],
                                                                                                 data_index[10]]
                        elif len(data_index[0]) == 10 and data_index[2] is not None:
                            page_data["Level3"]["data"].append(data_detail)

                        # 组织四级
                        elif len(data_index[0]) == 13 and data_index[2] is None:
                            page_data["Level4"]["code"].append(data_index[0] + "-" + data_index[1])
                            page_data["Level4"]["total"][data_index[0] + "-" + data_index[1]] = [data_index[8],
                                                                                                 data_index[10]]
                        elif len(data_index[0]) == 13 and data_index[2] is not None:
                            page_data["Level4"]["data"].append(data_detail)

                        # 组织五级
                        elif len(data_index[0]) == 16 and data_index[2] is None:
                            page_data["Level5"]["code"].append(data_index[0] + "-" + data_index[1])
                            page_data["Level5"]["total"][data_index[0] + "-" + data_index[1]] = [data_index[8],
                                                                                                 data_index[10]]
                        elif len(data_index[0]) == 16 and data_index[2] is not None:
                            page_data["Level5"]["data"].append(data_detail)
                    else:
                        continue
        return page_data

    def get_time_data(self):
        """
        获取表单中存在的有效数据，方便后续筛选的时候使用
        :return:self.time_data
        """
        self.time_data = set()
        for company in self.company_sheet_detail.keys():
            for company_date in self.company_sheet_detail[company].keys():
                self.time_data.add(company_date[-4:])
        self.time_data = list(self.time_data)
        self.time_data.sort()
        return self.time_data

    def calculate_year_basis(self):
        """
        计算环比，同比，并替换到原有数据中
        :return:
        """

        # 等级名字
        level_list = ["Level1", "Level2", "Level3", "Level4", "Level5"]

        # 处理所有数据
        for company_index in self.company_sheet_detail:  # 公司
            for date_index in self.company_sheet_detail[company_index]:  # 日期表
                for level_meg in level_list:  # 每个等级的所有数据都需要处理
                    if self.company_sheet_detail[company_index][date_index][level_meg]["data"]:  # 有数据
                        # 每一个数据行
                        for singe_item in self.company_sheet_detail[company_index][date_index][level_meg]["data"]:
                            data_index = self.company_sheet_detail[company_index][date_index][level_meg]["data"].index(
                                singe_item)  # 获取一下对应的数据行在数据中的位置，方便后面整体更换

                            # 当月同
                            num_1 = singe_item[3]
                            num_2 = self.get_target_data(data_dict=self.company_sheet_detail,
                                                         company=company_index,
                                                         date=self.get_compare_date(date_index)[1],
                                                         code=singe_item[0],
                                                         depart=singe_item[1],
                                                         )
                            # 当月环
                            num_3 = self.get_target_data(data_dict=self.company_sheet_detail,
                                                         company=company_index,
                                                         date=self.get_compare_date(date_index)[0],
                                                         code=singe_item[0],
                                                         depart=singe_item[1]
                                                         , )

                            if num_2 is None or float(num_2[3]) == 0.0:  # 没有数据，或者数据为0
                                tong_data = "0"
                            else:
                                tong_data = "{:.2f}%".format(((float(num_1) - float(num_2[3])) / float(num_2[3]))*100)

                            if num_3 is None or float(num_3[3]) == 0.0:
                                circle_data = "0"
                            else:
                                circle_data = "{:.2f}%".format(((float(num_1) - float(num_3[3])) / float(num_3[3]))*100)

                            # 当年同
                            num_a = singe_item[4]
                            num_b = self.get_target_data(data_dict=self.company_sheet_detail,
                                                         company=company_index,
                                                         date=self.get_compare_date(date_index)[1],
                                                         code=singe_item[0],
                                                         depart=singe_item[1],
                                                         )
                            if num_b is None or float(num_b[4]) == 0.0:
                                year_tong_data = "0"
                            else:
                                year_tong_data = "{:.2f}%".format(((float(num_a) - float(num_b[4])) / float(num_b[4]))*100)

                            # 组织新数据
                            new_singe_item = singe_item[:]
                            new_singe_item.append(tong_data)
                            new_singe_item.append(circle_data)
                            new_singe_item.append(year_tong_data)
                            # 替换原有数据
                            self.company_sheet_detail[company_index][date_index][level_meg]["data"][data_index] = new_singe_item

    def get_compare_date(self, data_string: str):
        """
        获取对应时间的同比、环比时间
        :param data_string: 日期数据解析（表单）
        :return: circle_bi, tong_bi  : 环比，同比
        """
        date_data = data_string[-4:]
        company_head = data_string[:-4]
        tong_bi = company_head + str(int(date_data) - 100)
        if date_data[-2:] == "01":
            circle_bi = company_head + str(int(date_data[0:2]) - 1) + str(12)
        else:
            circle_bi = company_head + str(int(date_data) - 1)
        return circle_bi, tong_bi

    def get_target_data(self, data_dict, company, date, code, depart):
        """
        获取对应环比、同比对应的数据行
        :param data_dict: 数据源
        :param company: 公司简写
        :param date: 公司+日期（表单）
        :param code: 科目代码
        :param depart: 部门
        :return:None / 对应数据行
        """
        # 匹配科目code，缩小搜索范围
        level = "Level1"
        if len(code) == 4:
            level = "Level1"
        elif len(code) == 7:
            level = "Level2"
        elif len(code) == 10:
            level = "Level3"
        elif len(code) == 13:
            level = "Level4"
        elif len(code) == 16:
            level = "Level5"

        # 是否有对应的表单
        try:
            data_temp = data_dict[company][date][level]["data"]
        except:
            return None

        # 如果有数据
        if data_temp:
            exist_flag = 0  # 没有对应行
            for item in data_temp:
                if item[0] == code and depart == item[1]:
                    exist_flag = 1
                    return item
            if exist_flag == 0:  # 完全没有数据
                return None


if __name__ == '__main__':
    a = FirstDeal("/Users/MikeImac/Desktop/FinanceTool/经营检测表（费用明细表）数据底稿.xlsx")
    print(a.company_sheet_detail)
